import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill, Font, Alignment
from functools import reduce
import matplotlib.pyplot as plt
from matplotlib.colors import to_hex, to_rgb
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.colors import to_hex, to_rgb, CSS4_COLORS, LinearSegmentedColormap, ListedColormap
from matplotlib.cm import ScalarMappable
from mycolorpy import colorlist as mcp
data = pd.read_excel("C:/Users/Анастасия/Downloads/analiz/975tcg integr_time 2/result with calibrate.xlsx")

plt.figure(figsize=(15, 6))
plt.subplots_adjust(wspace=0.5)
plt.subplots_adjust(hspace=0.5)

plt.subplot(1, 3, 1)
plt.title('Диод')
plt.xlabel('Диод')
plt.ylabel('Количество')
diode_counts, diode_bins, diode_info = plt.hist(data['diode'], color='#800000', alpha=0.5)
plt.hist(data['diode'], color='#800000', fill=False)

for rect, label in zip(diode_info, diode_counts):
    height = rect.get_height()
    if label != 0:
        plt.text(rect.get_x() + rect.get_width() / 2, height + 0.01, int(label),
                ha='center', va='bottom')

plt.subplot(1, 3, 2)
plt.title('Ток по спектру')
plt.xlabel('Ток по спектру')
plt.ylabel('Количество')
Id_int_counts, Id_int_bins, Id_int_info = plt.hist(data['Id_int'], color='#800000', fill=False)
plt.hist(data['Id_int'], color='#800000', alpha=0.5)

for rect, label in zip(Id_int_info, Id_int_counts):
    height = rect.get_height()
    if label != 0:
        plt.text(rect.get_x() + rect.get_width() / 2, height + 0.01, int(label),
                ha='center', va='bottom')

plt.subplot(1, 3, 3)
plt.title('Напряжение включения')
plt.xlabel('Напряжение включения')
plt.ylabel('Количество')
Vd_cur_counts, Vd_cur_bins, Vd_cur_info = plt.hist(data['Vd_cur'], bins=10, color='#800000', fill=False)
plt.hist(data['Vd_cur'], bins=10, color='#800000', alpha=0.5)

for rect, label in zip(Vd_cur_info, Vd_cur_counts):
    height = rect.get_height()
    if label != 0:
        plt.text(rect.get_x() + rect.get_width() / 2, height + 0.01, int(label),
                ha='center', va='bottom')
plt.savefig('diagrams.png')


plt.figure(figsize=(8, 6))
plt.title('FoM')
plt.xlabel('FoM')
plt.ylabel('Количество')

FoM_pow_counts, FoM_pow_bins, FoM_pow_info = plt.hist(data['FoM'], bins=20)

colors_hash = mcp.gen_color(cmap="Spectral_r", n=len(FoM_pow_bins)-1)
colors = []
for i in colors_hash:
    colors.append(i[1:])

for i in range(len(FoM_pow_bins)-1):
    FoM_pow_info[i].set_facecolor(colors_hash[i])

for rect, label in zip(FoM_pow_info, FoM_pow_counts):
    height = rect.get_height()
    if label != 0:
        plt.text(rect.get_x() + rect.get_width() / 2, height + 0.01, int(label),
                ha='center', va='bottom')

FoM_pow_max_index = []
for i in range(len(FoM_pow_bins)-1):
    if FoM_pow_counts[i] == np.max(FoM_pow_counts):
        FoM_pow_max_index.append(i)
        break

FoM_pow_middle_index = []
for i in range(len(FoM_pow_bins)-1):
    if 0.7*FoM_pow_bins[FoM_pow_max_index[0]] <= FoM_pow_bins[i] <= 2.3*FoM_pow_bins[FoM_pow_max_index[0]]:
        FoM_pow_middle_index.append(i)

for i in FoM_pow_middle_index:
    FoM_pow_info[i].set_hatch('o')

plt.legend([f'Всего дисплеев: {int(np.sum(FoM_pow_counts))} шт, 100%'
            f'\nштрих "o" - средняя партия: {int(np.sum(FoM_pow_counts[FoM_pow_middle_index]))}шт, {round(np.sum(FoM_pow_counts[FoM_pow_middle_index])/np.sum(FoM_pow_counts)*100, 1)} %'
            f'\nслева без "o" - брак: {int(np.sum(FoM_pow_counts[0:FoM_pow_middle_index[0]]))}шт, {round(np.sum(FoM_pow_counts[0:FoM_pow_middle_index[0]])/np.sum(FoM_pow_counts)*100, 1)} %'
            f'\nсправа без "o" - лучшие: {int(np.sum(FoM_pow_counts[FoM_pow_middle_index[-1]+1:]))}шт, {round(np.sum(FoM_pow_counts[FoM_pow_middle_index[-1]+1:])/np.sum(FoM_pow_counts)*100, 1)} %'
            ])
plt.savefig('FoM.png')
# plt.show()

display_map = {
    'coordinates': [],
    'FoM_low_cur': [],
    'FoM_big_cur': [],
    'diode': [],
    'Id_int': [],
    'Id_cur': [],
    'Vd_int': [],
    'Vd_cur': []
}

workbook = openpyxl.load_workbook(
    "C:/Users/Анастасия/Downloads/analiz/975tcg integr_time 2/result with calibrate 1.xlsx")
sheet = workbook.active
img1 = Image('diagrams.png')
sheet.add_image(img1, 'X1')
img2 = Image('FoM.png')
sheet.add_image(img2, 'X33')

eng = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'

Id_int_cell = sheet['L1']
Id_int_cell.value = '№'
Id_int_cell.font = Font(bold=True)

Id_int_cell = sheet['L4']
Id_int_cell.value = 'FoM'
Id_int_cell.font = Font(bold=True)

Id_int_cell = sheet['L5']
Id_int_cell.value = 'Диод'
Id_int_cell.font = Font(bold=True)

Id_int_cell = sheet['L6']
Id_int_cell.value = 'Напр вкл, В'
Id_int_cell.font = Font(bold=True)

Id_int_cell = sheet['L7']
Id_int_cell.value = 'Ток упр, мА'
Id_int_cell.font = Font(bold=True)

thick_up_border = Border(left=Side(style='thick'),
                         right=Side(style='thick'),
                         top=Side(style='thick'))

thick_mid_border = Border(left=Side(style='thick'),
                          right=Side(style='thick'))

thick_down_border = Border(left=Side(style='thick'),
                           right=Side(style='thick'),
                           bottom=Side(style='thick'))

for i in range(len(data)):
    length = 4
    x_coordinate = []
    if data['name'][i][1] == 'A' or data['name'][i][1] == 'А':
        x_coordinate.append(10)
    if data['name'][i][1] == 'B' or data['name'][i][1] == 'В':
        x_coordinate.append(11)
    if data['name'][i][1] == 'C' or data['name'][i][1] == 'С':
        x_coordinate.append(12)
    if len(x_coordinate) == 0:
        x_coordinate.append(int(data['name'][i][1]))

    x_coordinate = x_coordinate[0]

    y_array = np.arange(0, 58, 1)
    y_coordinate = y_array[- int(data['name'][i][0]) - length * (int(data['name'][i][0]) - 1)]
    # y_coordinate = y_array[- int(data['name'][i][0]) + length * (int(data['name'][i][0]) - 1)]


    name_cell = sheet[str(eng[22 - int(x_coordinate)]) + str(y_coordinate)]
    name_cell.value = data['name'][i]
    name_cell.border = thick_up_border
    name_cell.alignment = Alignment(horizontal='center')

    FoM_pow_cell = sheet[str(eng[22 - int(x_coordinate)]) + str(y_coordinate + 1)]
    FoM_pow_cell.value = data['FoM'][i]
    FoM_pow_cell.border = thick_mid_border
    FoM_pow_cell.alignment = Alignment(horizontal='center')

    diode_cell = sheet[str(eng[22 - int(x_coordinate)]) + str(y_coordinate + 2)]
    diode_cell.value = data['diode'][i]
    diode_cell.border = thick_mid_border
    diode_cell.alignment = Alignment(horizontal='center')

    Vd_cur_cell = sheet[str(eng[22 - int(x_coordinate)]) + str(y_coordinate + 3)]
    Vd_cur_cell.value = data['Vd_cur'][i]
    Vd_cur_cell.border = thick_mid_border
    Vd_cur_cell.alignment = Alignment(horizontal='center')

    Id_int_cell = sheet[str(eng[22 - int(x_coordinate)]) + str(y_coordinate + 4)]
    Id_int_cell.value = data['Id_int'][i]
    Id_int_cell.border = thick_down_border
    Id_int_cell.alignment = Alignment(horizontal='center')

    for a in range(len(FoM_pow_bins) - 1):
        if FoM_pow_bins[a] <= data['FoM'][i] <= FoM_pow_bins[a + 1]:
            name_cell.fill = PatternFill('solid', fgColor=colors[a])

workbook.save("C:/Users/Анастасия/Downloads/analiz/975tcg integr_time 2/result with calibrate 1.xlsx")